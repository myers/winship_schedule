#!/usr/bin/env python3
 
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import take2
import winship_schedule
from datetime import timedelta

def get_iso_week(d):
    # If the date is not a Sunday, get the previous Sunday
    while d.weekday() != 6:  # 6 represents Sunday
        d -= timedelta(days=1)
    return d.isocalendar()[1]

def get_colors(share):
    # Generate unique colors for each share (background, font)
    colors = {
        'becca': ('B2B2B2', '000000'),
        'david': ('287289', 'FFFFFF'),
        'hugh': ('FFFFC1', '000000'),
        'eddie': ('2A4C7F', 'FFFFFF'),
        'frank_latimer': ('F7B17D', '000000'),
        'frank_may': ('B8B085', '000000'),
        'hankey': ('AAC0DE', '000000'),
        'jim': ('17A43F', 'FFFFFF'),
        'joe': ('C2FFC0', '000000'),
        'myers': ('FC4C06', 'FFFFFF'),
        'lane': ('ead203', '000000'),
        'hayley': ('AF2488', 'FFFFFF'),
        'jordan': ('5483FF', '000000'),
        'richard': ('A56193', 'FFFFFF'),
        'will': ('FFA1A2', '000000')
    }
    return colors.get(share.split('-')[0], ('FFFFFF', '000000'))

def export_to_excel(filename, schedule):
    # Determine start and end years from schedule
    start_year = min(hy.year for hy in schedule)
    end_year = max(hy.year for hy in schedule)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Winship House Schedule"

    # Set up header row with years
    ws.cell(row=1, column=1, value="Week")
    for year in range(start_year, end_year + 1):
        ws.cell(row=1, column=(year - start_year + 2), value=year)

    # Fill in the schedule
    print(f"start_year: {start_year} end_year: {end_year}")
    for year in range(start_year, end_year + 1):
        print(f"year: {year}")
        house_year = next((hy for hy in schedule if hy.year == year), None)
        if not house_year:
            print(f"Warning: No schedule found for year {year}")
            continue
            
        column = year - start_year + 2
        
        for week in house_year.weeks:
            print(f"week: {week.start} {week.end} {week.kind} {week.share}")
            iso_week = get_iso_week(week.start)
            if iso_week <= 9:
                continue  # Skip the first 9 weeks
            row = iso_week - 9 + 1  # +1 because row 1 is the header, -9 to adjust for skipped weeks
            cell = ws.cell(row=row, column=column)
            share_name = winship_schedule.share_name_to_name(week.share)
            
            # Add emoji if it's a holiday week
            if week.holiday:
                emoji = take2.holiday_to_emoji(week.holiday)
                cell.value = f"{share_name} {emoji}"
            else:
                cell.value = share_name
                
            bg_color, font_color = get_colors(week.share)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.font = Font(color=font_color)
            
            # Add comment with date range, chunk type, and holiday (if any)
            date_range = f"{week.start.strftime('%Y-%m-%d')}"
            comment_text = f"{date_range}\n{week.kind}"
            if hasattr(week, 'holiday') and week.holiday:
                comment_text += f"\nHoliday: {week.holiday}"
            comment = Comment(comment_text, "Winship Schedule")
            cell.comment = comment

        # Fill in the week numbers for this year
        if year == start_year:
            for week_num in range(10, 54):
                ws.cell(row=week_num - 9 + 1, column=1, value=f"Week {week_num}")

    # Adjust column widths
    for column in range(1, end_year - start_year + 3):
        ws.column_dimensions[get_column_letter(column)].width = 15

    # Adjust row height
    for row in range(1, 46):  # 44 weeks (53-9) + header
        ws.row_dimensions[row].height = 30

    wb.save(filename)

if __name__ == "__main__":
    export_to_excel("winship_schedule.xlsx", schedule)
